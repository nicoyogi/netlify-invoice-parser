import fitz  # PyMuPDF
import re
import pandas as pd
import os
import io
import json
import base64
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def parse_amount(amount_str):
    """Mengonversi string angka gaya Eropa (mis., '1.234,56') menjadi float."""
    if not isinstance(amount_str, str):
        return 0.0
    cleaned = amount_str.replace('.', '').replace(',', '.')
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0

def process_pdf_to_excel(pdf_content, filename):
    """Fungsi utama untuk mengekstrak data dari konten PDF dan menghasilkan file Excel."""
    print("Mulai memproses PDF...")
    doc = fitz.open(stream=pdf_content, filetype="pdf")
    text = ""
    for page in doc:
        text += page.get_text()
    
    # DEBUG: Log bagian awal teks yang diekstrak untuk diagnostik
    print(f"--- Teks yang Diekstrak (500 karakter pertama) ---\n{text[:500]}\n-----------------------------------------")
    
    def find(pattern, source_text=text):
        """Fungsi bantu yang aman untuk regex."""
        match = re.search(pattern, source_text, re.DOTALL | re.IGNORECASE)
        return match.group(1).strip() if match else None

    # --- Ekstraksi Data Header ---
    print("Mengekstrak data header...")
    invoice_number = find(r'Rechnungs Nr\.:\s*(\d+)') or 'N/A'
    sender = find(r'Absender:\s*([^\n]+)') or 'N/A'
    etd_eta = find(r'ETD/ETA:\s*([^\n]+)') or 'N/A'
    port_loading = find(r'Port of Loading:\s*([^\n]+)') or 'N/A'
    port_discharge = find(r'Port of Discharge:\s*([^\n]+)') or 'N/A'
    invoice_date = find(r'Rechnungsdatum:\s*(\d{2}-[A-Za-z]{3}-\d{4})') or 'N/A'
    stt_number = find(r'STT Nr\.:\s*(\d+)') or 'N/A'
    gross_weight_kg = find(r'Bruttogewicht\s*([\d.,]+)\s*KGS') or '0'
    volume_cbm = find(r'Volumen\s*([\d.,]+)\s*CBM') or '0'
    print(f"Invoice: {invoice_number}, Date: {invoice_date}")

    # --- Ekstraksi Rincian Biaya ---
    print("Mencari blok biaya...")
    # PERBAIKAN UTAMA: Menggunakan anchor yang lebih stabil dan fleksibel ('Gesamtbetrag')
    # untuk menandai akhir dari blok biaya. Ini lebih tahan terhadap variasi format.
    cost_section = find(r"Unsere Leistungen(.*?)\s*Gesamtbetrag")
    
    if not cost_section:
        print("ERROR: Blok biaya 'Unsere Leistungen' tidak ditemukan. Regex gagal.")
        # Log seluruh teks jika regex gagal, ini sangat membantu untuk debugging.
        print(f"--- TEKS LENGKAP UNTUK DEBUGGING ---\n{text}\n---------------------------------")
        raise ValueError("Tidak dapat menemukan blok rincian biaya ('Unsere Leistungen') dalam PDF.")

    print("Blok biaya ditemukan. Mengekstrak setiap item biaya...")
    rows = []
    cost_label_map = {
        "Summarische Eingangsmeldung": "ENS",
        "Seefracht": "SFRT",
        r"THC \(Terminal Handling Charge\)": "THC",
        "Abfertigungskosten im": "CCDE",
        r"ISPS \(Hafen & Terminal": "ISPS",
        "Nachlaufkosten": "NL",
        "Delivery-/Drop-Off-Gebühr": "DROP",
        "Importverzollung in NL": "Zoll"
    }

    for label_pattern, code in cost_label_map.items():
        amount_str = find(rf"{label_pattern}.*?EUR\s+([\d.,]+)", source_text=cost_section)
        
        if amount_str:
            amount_float = parse_amount(amount_str)
            print(f"Ditemukan: {code} = {amount_float}")
            rows.append({
                "file": filename, "invoice_number": invoice_number, "sender": sender,
                "etd_eta": etd_eta, "port_loading": port_loading, "port_discharge": port_discharge,
                "invoice_date": invoice_date, "stt_number": stt_number,
                "gross_weight_kg": gross_weight_kg, "volume_cbm": volume_cbm,
                "cost_type": code, "amount": amount_float
            })

    if not rows:
        print("ERROR: Tidak ada baris biaya yang berhasil diekstrak.")
        raise ValueError("Tidak ada rincian biaya yang dapat diekstrak. Format PDF mungkin berbeda.")

    # --- Pembuatan File Excel di Memori ---
    print(f"Berhasil mengekstrak {len(rows)} baris biaya. Membuat file Excel...")
    df_long = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_long.to_excel(writer, index=False, sheet_name='InvoiceData')
        ws = writer.sheets['InvoiceData']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws.column_dimensions[col_letter].width = max_length + 4
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            cell = row[-1]
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print("Pembuatan Excel selesai.")
    return output.getvalue()

def handler(event, context):
    """Fungsi handler yang dipanggil oleh Netlify saat ada request."""
    try:
        print("Fungsi handler dipanggil.")
        
        if not event.get('isBase64Encoded'):
             raise ValueError("Body request tidak ter-encode base64.")

        filename = event['headers'].get('x-filename', 'unknown.pdf')
        print(f"Menerima file: {filename}")
        
        pdf_content = base64.b64decode(event['body'])
        print(f"Berhasil mem-parsing konten PDF dengan panjang: {len(pdf_content)} bytes.")

        excel_data = process_pdf_to_excel(pdf_content, filename)

        print("Mengirimkan file Excel sebagai respons.")
        return {
            "statusCode": 200,
            "headers": {
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Disposition": f"attachment; filename=\"parsed_{os.path.splitext(filename)[0]}.xlsx\""
            },
            "body": base64.b64encode(excel_data).decode('utf-8'),
            "isBase64Encoded": True
        }
    except Exception as e:
        tb_str = traceback.format_exc()
        error_message_for_log = f"Terjadi kesalahan fatal di server: {type(e).__name__} - {e}\nTraceback:\n{tb_str}"
        print(f"ERROR: {error_message_for_log}")
        
        user_error_message = f"Terjadi kesalahan di server: {type(e).__name__} - {str(e)}"
        return {
            "statusCode": 500,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"error": user_error_message})
        }
